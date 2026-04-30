"""
VICSA Conectados - Scraper RAI SPA v13
========================================
Cambios respecto a v12:
- MODO RÁPIDO  (--modo rapido):  solo stock/precio de SKUs conocidos → ~3 min
  Corre cada 1 hora en GitHub Actions. Lee el último JSON completo y visita
  solo los listados para detectar cambios de disponibilidad y precio.
  Salida: stock_delta_FECHA.json con solo los SKUs que cambiaron.

- MODO COMPLETO (--modo completo, default): catálogo completo con fichas → ~30 min
  Corre 1 vez al día (madrugada). Detecta productos nuevos, actualiza imágenes
  y especificaciones.
  Salida: catalogo_vicsa_FECHA.json (igual que v12)

- DEDUPLICACIÓN de SKUs dentro del mismo producto base (fix bug v12)
- Stock conservador: el JSON incluye campo 'stock_publicar' = 50% del stock
  detectado, con mínimo 1 y máximo 10, para publicar en ML con margen de seguridad

Uso local:
    $env:VICSA_USER = "email@ejemplo.cl"
    $env:VICSA_PASS = "tu_contraseña"

    # Modo completo (default, ~30 min)
    python scraper_v13.py --categorias "Calzado de Seguridad" "Ropa Tecnica" "Ropa de Trabajo"

    # Modo rápido (~3 min) — requiere un JSON completo previo en output_vicsa/
    python scraper_v13.py --modo rapido

    # Modo rápido para categorías específicas
    python scraper_v13.py --modo rapido --categorias "Calzado de Seguridad"

GitHub Actions:
    Modo rápido:   cron '0 * * * *'    (cada hora)
    Modo completo: cron '0 3 * * *'    (cada día a las 3am)
"""

import asyncio, csv, re, sys, json, argparse, os
from pathlib import Path
from datetime import datetime

try:
    from playwright.async_api import async_playwright
except ImportError:
    print("pip install playwright && playwright install chromium"); sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    TIENE_XLSX = True
except ImportError:
    TIENE_XLSX = False

# ─── URLs ────────────────────────────────────────────────────────────────────
BASE      = "https://www.vicsaconectados.cl/cgi-bin/procesa.pl"
IMG_BASE  = "https://www.vicsaconectados.cl/catalogo/images/sys"
LOGIN_URL = BASE + "?plantilla=/vicsamobile/v2/login.html"
LIST_URL  = BASE + "?plantilla=/vicsamobile/v2/listado_producto.html&id_cat={id_cat}"
PAGE_URL  = BASE + "?plantilla=/vicsamobile/v2/listado_producto.html&id_cat={id_cat}&p={p}"
PROD_URL  = BASE + "?plantilla=/vicsamobile/v2/producto.html&id_cat={id_cat}&id_prod={id_prod}"

OUTPUT_DIR = Path("output_vicsa")
IMG_DIR    = OUTPUT_DIR / "imagenes"
DELAY      = 0.6  # ligeramente más agresivo que v12, Vicsa lo tolera

# ─── Categorías ──────────────────────────────────────────────────────────────
CATEGORIAS = {
    "192": "Cascos", "771": "Cascos", "772": "Cascos", "773": "Cascos",
    "1003": "Cascos", "1004": "Cascos", "1026": "Cascos", "1031": "Cascos",
    "1067": "Cascos", "1068": "Cascos", "1166": "Cascos", "1190": "Cascos",
    "1191": "Cascos", "1203": "Cascos", "1204": "Cascos", "1205": "Cascos",
    "1206": "Cascos", "1207": "Cascos", "1208": "Cascos", "1209": "Cascos",
    "1210": "Cascos",
    "143": "Lentes de Seguridad", "144": "Lentes de Seguridad", "145": "Lentes de Seguridad",
    "146": "Lentes de Seguridad", "147": "Lentes de Seguridad", "148": "Lentes de Seguridad",
    "149": "Lentes de Seguridad", "150": "Lentes de Seguridad", "151": "Lentes de Seguridad",
    "152": "Lentes de Seguridad", "153": "Lentes de Seguridad", "154": "Lentes de Seguridad",
    "155": "Lentes de Seguridad", "156": "Lentes de Seguridad", "157": "Lentes de Seguridad",
    "159": "Lentes de Seguridad", "160": "Lentes de Seguridad", "169": "Lentes de Seguridad",
    "411": "Lentes de Seguridad", "589": "Lentes de Seguridad", "774": "Lentes de Seguridad",
    "776": "Lentes de Seguridad", "777": "Lentes de Seguridad", "1037": "Lentes de Seguridad",
    "1038": "Lentes de Seguridad", "1039": "Lentes de Seguridad", "1040": "Lentes de Seguridad",
    "1050": "Lentes de Seguridad",
    "253": "Guantes", "265": "Guantes", "266": "Guantes", "267": "Guantes",
    "268": "Guantes", "295": "Guantes", "299": "Guantes", "306": "Guantes",
    "457": "Guantes", "466": "Guantes", "468": "Guantes", "469": "Guantes",
    "572": "Guantes", "573": "Guantes", "575": "Guantes", "576": "Guantes",
    "585": "Guantes", "586": "Guantes", "765": "Guantes", "768": "Guantes",
    "769": "Guantes", "794": "Guantes", "795": "Guantes", "796": "Guantes",
    "798": "Guantes", "806": "Guantes", "1073": "Guantes", "1076": "Guantes",
    "1078": "Guantes", "1088": "Guantes", "1089": "Guantes", "1090": "Guantes",
    "1099": "Guantes", "1102": "Guantes", "1103": "Guantes",
    "172": "Proteccion Auditiva", "174": "Proteccion Auditiva", "175": "Proteccion Auditiva",
    "591": "Proteccion Auditiva", "620": "Proteccion Auditiva", "629": "Proteccion Auditiva",
    "178": "Proteccion Respiratoria", "202": "Proteccion Respiratoria", "204": "Proteccion Respiratoria",
    "757": "Proteccion Respiratoria", "759": "Proteccion Respiratoria", "1070": "Proteccion Respiratoria",
    "1192": "Proteccion Respiratoria", "1193": "Proteccion Respiratoria", "1194": "Proteccion Respiratoria",
    "1196": "Proteccion Respiratoria", "1198": "Proteccion Respiratoria", "1199": "Proteccion Respiratoria",
    "1201": "Proteccion Respiratoria",
    "176": "Proteccion Facial", "208": "Proteccion Facial", "209": "Proteccion Facial",
    "223": "Proteccion Facial", "224": "Proteccion Facial", "1035": "Proteccion Facial",
    "987": "Calzado de Seguridad", "995": "Calzado de Seguridad", "1064": "Calzado de Seguridad",
    "1065": "Calzado de Seguridad", "1104": "Calzado de Seguridad", "1106": "Calzado de Seguridad",
    "1107": "Calzado de Seguridad", "1110": "Calzado de Seguridad", "1115": "Calzado de Seguridad",
    "1116": "Calzado de Seguridad", "1117": "Calzado de Seguridad", "1118": "Calzado de Seguridad",
    "1120": "Calzado de Seguridad", "1125": "Calzado de Seguridad", "1126": "Calzado de Seguridad",
    "1135": "Calzado de Seguridad", "1141": "Calzado de Seguridad", "1142": "Calzado de Seguridad",
    "1144": "Calzado de Seguridad", "1155": "Calzado de Seguridad", "1159": "Calzado de Seguridad",
    "1161": "Calzado de Seguridad",
    "181": "Ropa de Trabajo", "183": "Ropa de Trabajo", "184": "Ropa de Trabajo",
    "200": "Ropa de Trabajo", "216": "Ropa de Trabajo", "217": "Ropa de Trabajo",
    "220": "Ropa de Trabajo", "225": "Ropa de Trabajo", "228": "Ropa de Trabajo",
    "836": "Ropa de Trabajo", "847": "Ropa de Trabajo", "851": "Ropa de Trabajo",
    "852": "Ropa de Trabajo", "853": "Ropa de Trabajo", "854": "Ropa de Trabajo",
    "856": "Ropa de Trabajo", "858": "Ropa de Trabajo", "1093": "Ropa de Trabajo",
    "1094": "Ropa de Trabajo", "1098": "Ropa de Trabajo",
    "197": "Ropa Tecnica", "199": "Ropa Tecnica", "284": "Ropa Tecnica",
    "287": "Ropa Tecnica", "288": "Ropa Tecnica", "289": "Ropa Tecnica",
    "740": "Ropa Tecnica", "741": "Ropa Tecnica", "778": "Ropa Tecnica",
    "779": "Ropa Tecnica", "955": "Ropa Tecnica", "957": "Ropa Tecnica",
    "958": "Ropa Tecnica",
    "191": "Seguridad Vial", "193": "Seguridad Vial", "195": "Seguridad Vial",
    "210": "Seguridad Vial", "211": "Seguridad Vial", "212": "Seguridad Vial",
    "214": "Seguridad Vial", "234": "Seguridad Vial", "239": "Seguridad Vial",
    "240": "Seguridad Vial", "305": "Seguridad Vial",
    "308": "Primeros Auxilios", "309": "Primeros Auxilios", "310": "Primeros Auxilios",
    "810": "Primeros Auxilios", "817": "Primeros Auxilios", "823": "Primeros Auxilios",
    "824": "Primeros Auxilios", "825": "Primeros Auxilios", "826": "Primeros Auxilios",
    "829": "Primeros Auxilios",
    "781": "Proteccion Solar", "782": "Proteccion Solar", "791": "Proteccion Solar",
    "792": "Proteccion Solar",
    "860": "Ergonomia",
}

COLORES = [
    "HIGH VISION","HIGT VISION","FLUOR VERDE","FLUOR AMARILLO",
    "AMARILLO","NARANJA","NARANJO","AZUL","BLANCO","NEGRO","ROJO",
    "VERDE","GRIS","CAFE","MARRON","BEIGE","CELESTE","MORADO","LILA",
    "ROSADO","ROSA","FUCSIA","FLUOR","PLATEADO","DORADO","TRANSPARENTE",
    "HUMO","SMOKE","CLEAR","AMBAR","MULTICOLOR","NATURAL","MIEL",
]

PATRON_TALLA = re.compile(
    r'\b(XXXL|XXL|XL|XS|XXS|(?<!\w)L(?!\w)|(?<!\w)M(?!\w)|(?<!\w)S(?!\w))\b'
    r'|T\.?\s*(\d{2})'
    r'|\bTALLA\s+([A-Z0-9]+)\b'
    r'|\b(3[4-9]|4[0-9]|5[0-2])\b(?=\s|$)'
    r'|\b(UNICA|ÚNICA)\b',
    re.IGNORECASE
)

# ─── Helpers ─────────────────────────────────────────────────────────────────
def detectar_color_talla(nombre):
    n = nombre.upper()
    color = ""
    for c in sorted(COLORES, key=len, reverse=True):
        if c in n:
            color = c.title()
            break
    talla = ""
    m = PATRON_TALLA.search(n)
    if m:
        talla = next((g for g in m.groups() if g), "").strip().upper()
        talla = re.sub(r'^T\.?\s*', 'T.', talla) if re.match(r'^\d{2}$', talla) else talla
        talla = re.sub(r'^TALLA\s+', '', talla)
    return color, talla

def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

def parse_precio(texto):
    limpio = re.sub(r'[^\d]', '', (texto or ''))
    try: return float(limpio)
    except: return 0.0

def limpiar(texto):
    if not texto: return ""
    reemplazos = {'ï¿½':'','Ã©':'é','Ã³':'ó','Ã¡':'á','Ã­':'í','Ã±':'ñ','\n':' ','\r':''}
    for k,v in reemplazos.items():
        texto = texto.replace(k,v)
    return re.sub(r'\s+', ' ', texto).strip()

def stock_conservador(stock_vicsa: int) -> int:
    """
    Convierte el stock real de Vicsa en el stock que publicamos en ML.
    Regla: 50% del stock de Vicsa, mínimo 1 si hay stock, máximo 10.
    Si Vicsa tiene 0 → publicamos 0 (pausar publicación en ML).
    """
    if stock_vicsa <= 0:
        return 0
    conservador = max(1, stock_vicsa // 2)
    return min(conservador, 10)

def ultimo_json_completo() -> Path | None:
    """Retorna el JSON completo más reciente en output_vicsa/"""
    jsons = sorted(
        [f for f in OUTPUT_DIR.glob("catalogo_vicsa_*.json")],
        key=lambda f: f.stat().st_mtime,
        reverse=True
    )
    return jsons[0] if jsons else None

# ─── Login automático ─────────────────────────────────────────────────────────
async def login_automatico(page):
    usuario  = os.environ.get("VICSA_USER", "").strip()
    password = os.environ.get("VICSA_PASS", "").strip()
    if not usuario or not password:
        print("\n" + "="*60)
        print("  ERROR: Variables de entorno no encontradas.")
        print("  PowerShell:")
        print("    $env:VICSA_USER = 'tu_usuario'")
        print("    $env:VICSA_PASS = 'tu_contraseña'")
        print("  Linux/Mac:")
        print("    export VICSA_USER='tu_usuario'")
        print("    export VICSA_PASS='tu_contraseña'")
        print("="*60 + "\n")
        sys.exit(1)

    log("Login automático...")
    await page.goto(LOGIN_URL, wait_until="domcontentloaded")
    await asyncio.sleep(1.5)
    html_login = await page.content()

    user_field = pass_field = None

    # Detectar campo usuario (incluye "email" — campo real de Vicsa)
    for pattern in [r'name="([^"]*(?:email|user|rut|login|cuenta)[^"]*)"', r'name="([^"]*RUT[^"]*)"']:
        m = re.search(pattern, html_login, re.IGNORECASE)
        if m:
            user_field = m.group(1)
            break

    # Detectar campo contraseña (incluye "clave" — campo real de Vicsa)
    for pattern in [r'name="([^"]*(?:clave|pass|password|contrase)[^"]*)"']:
        m = re.search(pattern, html_login, re.IGNORECASE)
        if m:
            pass_field = m.group(1)
            break

    # Fallback por tipo de input
    if not user_field:
        inputs = re.findall(r'<input[^>]*type=["\']?(?:text|email)["\']?[^>]*name=["\']([^"\']+)["\']', html_login, re.IGNORECASE)
        if inputs: user_field = inputs[0]
    if not pass_field:
        inputs = re.findall(r'<input[^>]*type=["\']?password["\']?[^>]*name=["\']([^"\']+)["\']', html_login, re.IGNORECASE)
        if inputs: pass_field = inputs[0]

    if not user_field or not pass_field:
        log("ERROR: No se detectaron campos de login.")
        all_inputs = re.findall(r'<input[^>]+>', html_login, re.IGNORECASE)
        for inp in all_inputs[:10]: log(f"  {inp[:120]}")
        sys.exit(1)

    log(f"Campos: usuario='{user_field}' | contraseña='{pass_field}'")
    await page.fill(f"input[name='{user_field}']", usuario)
    await asyncio.sleep(0.3)
    await page.fill(f"input[name='{pass_field}']", password)
    await asyncio.sleep(0.3)

    submitted = False
    for selector in ["input[type='submit']", "button[type='submit']",
                     "button:has-text('Ingresar')", "button:has-text('Acceder')",
                     "input[value='Acceder']", "input[value='Ingresar']"]:
        try:
            await page.click(selector, timeout=2000)
            submitted = True
            break
        except: continue

    if not submitted:
        await page.press(f"input[name='{pass_field}']", "Enter")

    await asyncio.sleep(3)

    url_actual = page.url
    html_post  = await page.content()

    if "login" in url_actual.lower() and "home" not in url_actual.lower():
        log(f"ERROR: Login falló. URL: {url_actual}")
        sys.exit(1)
    if any(x in html_post.lower() for x in ["contraseña incorrecta","usuario incorrecto",
                                              "error de acceso","clave incorrecta","invalid"]):
        log("ERROR: Credenciales incorrectas.")
        sys.exit(1)

    log(f"Login OK ✓ → {url_actual}")

# ─── Extracción del listado ───────────────────────────────────────────────────
def extraer_tallas_html(html: str) -> dict:
    """Mapea item_N → talla extraída del <td> visible 'talla XX' que sigue al codExt."""
    tallas = {}
    for m in re.finditer(r'name="item_(\d+)_prod_1_codExt"[^>]*value="\d+"', html):
        n   = m.group(1)
        end = m.end()
        ventana = html[end:end+4000]
        talla_m = re.search(
            r'<span[^>]*>\s*Talla\s*</span>\s*talla\s+([A-Z0-9./\-]+)',
            ventana, re.IGNORECASE
        )
        if talla_m:
            t = talla_m.group(1).strip().upper()
            t = re.sub(r'^T\.?', '', t)
            tallas[n] = t
    return tallas

def extraer_stock_html(html: str) -> dict:
    """Mapea item_N → stock real (Pud) por variante.
    Vicsa expone 'Pud N' en cada fila de la tabla y un botón 'Sin Stock' si N=0."""
    stocks = {}
    for m in re.finditer(r'name="item_(\d+)_prod_1_codExt"[^>]*value="\d+"', html):
        n   = m.group(1)
        end = m.end()
        ventana = html[end:end+4000]
        # Cortar la ventana al siguiente item para no leer stock del vecino
        sig = re.search(r'name="item_\d+_prod_1_codExt"', ventana)
        if sig:
            ventana = ventana[:sig.start()]
        pud_m = re.search(r'<span[^>]*>\s*Pud\s*</span>\s*(\d+)', ventana, re.IGNORECASE)
        if pud_m:
            stocks[n] = int(pud_m.group(1))
        elif re.search(r'value="Sin Stock"', ventana, re.IGNORECASE):
            stocks[n] = 0
    return stocks

def extraer_productos_listado(html, cat_nombre):
    idprods = dict(re.findall(r'name="item_(\d+)_prod_1_idprod"[^>]*value="(\d+)"', html))
    codexts = dict(re.findall(r'name="item_(\d+)_prod_1_codExt"[^>]*value="([^"]+)"', html))
    nombres = dict(re.findall(r'name="msgNombre(\d+)"[^>]*value="([^"]+)"', html))
    precios = dict(re.findall(r'name="msgPrecio(\d+)"[^>]*value="([\d\.\,]+)"', html))
    tallas_html = extraer_tallas_html(html)
    urls_cat = re.findall(r'id_cat=(\d+)&(?:amp;)?id_prod=(\d+)', html)
    idprod_to_cat = {idp: idc for idc, idp in urls_cat}

    # Detectar stock desde el listado
    # Vicsa muestra disponibilidad con clases CSS o texto "Sin Stock" / "Disponible"
    stock_map = {}
    for n in idprods.keys():
        # Buscar indicador de stock por item (heurística basada en HTML de Vicsa)
        bloque = re.search(
            rf'item_{n}_prod_1_idprod.{{0,2000}}?(?:sinStock|sin.stock|agotado|no.disponible)',
            html, re.IGNORECASE | re.DOTALL
        )
        stock_map[n] = 0 if bloque else 5  # Si encuentra "sin stock" → 0, sino → 5 (conservador)

    productos = []
    vistos_sku = set()

    for n in sorted(idprods.keys(), key=int):
        sku     = codexts.get(n, "").strip()
        id_prod = idprods.get(n, "")
        nombre  = limpiar(nombres.get(n, ""))
        precio  = parse_precio(precios.get(n, "0"))
        id_cat  = idprod_to_cat.get(id_prod, "0")
        stock_v = stock_map.get(n, 5)

        if not sku or not nombre or sku in vistos_sku:
            continue
        vistos_sku.add(sku)

        color, talla = detectar_color_talla(nombre)
        if not talla:
            talla = tallas_html.get(n, "")

        productos.append({
            "sku":              sku,
            "nombre":           nombre,
            "categoria":        cat_nombre,
            "talla":            talla,
            "color":            color,
            "precio_neto":      precio,
            "stock_vicsa":      stock_v,
            "stock_publicar":   stock_conservador(stock_v),
            "especificaciones": "",
            "ficha_tecnica_url":"",
            "certificaciones":  "",
            "imagen":           "",
            "url_producto":     PROD_URL.format(id_cat=id_cat, id_prod=id_prod),
            "_id_prod":         id_prod,
            "_id_cat":          id_cat,
            "_actualizado":     datetime.now().isoformat(),
        })

    return productos

async def enriquecer_producto(page, prod, descargar_img):
    """Visita la ficha individual para obtener talla, specs, imagen. Solo modo completo."""
    try:
        await page.goto(prod["url_producto"], wait_until="domcontentloaded")
        await asyncio.sleep(DELAY * 0.8)
        html = await page.content()

        nombre_m = re.search(r'class="productoNombre"><span>([^<]+)', html)
        if nombre_m:
            nombre = limpiar(nombre_m.group(1))
            prod["nombre"] = nombre
            color, talla = detectar_color_talla(nombre)
            if color: prod["color"] = color
            if talla: prod["talla"] = talla

        # Talla extraída del <td> visible que coincide con el SKU de este producto.
        # Localiza el codExt del SKU y busca solo en los siguientes 2KB para no cruzar
        # al item siguiente.
        if not prod.get("talla"):
            sku_actual = prod.get("sku", "")
            if sku_actual:
                idx = html.find(f'value="{sku_actual}"')
                if idx >= 0:
                    ventana = html[idx:idx+2500]
                    m = re.search(
                        r'<span[^>]*>\s*Talla\s*</span>\s*talla\s+([A-Z0-9./\-]+)',
                        ventana, re.IGNORECASE
                    )
                    if m:
                        t = m.group(1).strip().upper()
                        t = re.sub(r'^T\.?', '', t)
                        prod["talla"] = t

        # NO sobrescribir el SKU: la ficha individual lista TODOS los SKUs del producto
        # (uno por talla), pero el SKU correcto para esta variante es el que vino del listado.
        # Sobrescribir colapsa todos los variantes a un único SKU y la dedup posterior
        # los elimina (bug histórico: Apollo Cafe 13 tallas → 1 sola variante).
        # cod_m = re.search(r'productoCodigo[^>]*>.*?(\d{8,})', html, re.DOTALL)
        # if cod_m:
        #     prod["sku"] = cod_m.group(1).strip()

        precio_m = re.search(r'Precio distribuidor[^\$]*\$([\d\.\,]+)', html)
        if precio_m:
            prod["precio_neto"] = parse_precio(precio_m.group(1))

        # Stock POR TALLA: extrae el Pud específico para el SKU de esta variante.
        # No usar un check global "sin stock" en toda la ficha — eso colapsa todas
        # las tallas a 0 si solo una está sin stock (bug histórico).
        sku_actual = prod.get("sku", "")
        if sku_actual:
            idx = html.find(f'value="{sku_actual}"')
            if idx >= 0:
                # Buscar el siguiente item_N para no cruzar al vecino
                ventana = html[idx:idx+4000]
                sig = re.search(r'name="item_\d+_prod_1_codExt"', ventana[100:])
                if sig:
                    ventana = ventana[:100 + sig.start()]
                pud_m = re.search(r'<span[^>]*>\s*Pud\s*</span>\s*(\d+)', ventana, re.IGNORECASE)
                if pud_m:
                    stock_ficha = int(pud_m.group(1))
                elif re.search(r'value="Sin Stock"', ventana, re.IGNORECASE):
                    stock_ficha = 0
                else:
                    stock_ficha = 5  # fallback conservador
                prod["stock_vicsa"]    = stock_ficha
                prod["stock_publicar"] = stock_conservador(stock_ficha)

        for aleta in re.findall(r'aleta[^"]*"[^"]*"[^>]*>(.*?)</(?:div|section)', html, re.DOTALL | re.IGNORECASE):
            txt = limpiar(re.sub(r'<[^>]+>', ' ', aleta))
            if not txt: continue
            al = txt.lower()
            if 'especificaci' in al and not prod["especificaciones"]:
                prod["especificaciones"] = txt[:500]
            elif ('ficha' in al or 'técnica' in al) and not prod["ficha_tecnica_url"]:
                arch_m = re.search(r'id_archivo=(\d+)', aleta)
                if arch_m:
                    prod["ficha_tecnica_url"] = BASE + f"?plantilla=/vicsamobile/v2/archivo.html&id_archivo={arch_m.group(1)}&download=1"
            elif 'certif' in al and not prod["certificaciones"]:
                prod["certificaciones"] = txt[:300]

        img_m = re.search(r'(\d+)_(\d+)_ficha\.(png|gif|jpg)', html, re.IGNORECASE)
        if img_m:
            fn = f"{img_m.group(1)}_{img_m.group(2)}_ficha.{img_m.group(3)}"
            prod["imagen"] = fn
            if descargar_img:
                img_path = IMG_DIR / fn
                if not img_path.exists():
                    try:
                        resp = await page.request.get(f"{IMG_BASE}/{fn}")
                        if resp.ok:
                            img_path.write_bytes(await resp.body())
                    except: pass
    except:
        pass
    return prod

async def scrape_categoria(page, id_cat, cat_nombre, limite, todos, vistos):
    p_num = 1
    max_paginas = 999
    skus_pagina_anterior = set()
    total_nuevos = 0

    while p_num <= max_paginas:
        url = LIST_URL.format(id_cat=id_cat) if p_num == 1 else PAGE_URL.format(id_cat=id_cat, p=p_num)
        await page.goto(url, wait_until="domcontentloaded")
        await asyncio.sleep(DELAY * 0.5)
        html = await page.content()

        if p_num == 1:
            pag_m = re.search(r'/ (\d+)</div>', html)
            max_paginas = int(pag_m.group(1)) if pag_m else 1

        prods = extraer_productos_listado(html, cat_nombre)
        if not prods: break

        skus_esta = {pr["sku"] for pr in prods}
        if p_num > 1 and skus_esta == skus_pagina_anterior: break
        skus_pagina_anterior = skus_esta

        for pr in prods:
            sku_key = pr["sku"] or pr["_id_prod"]
            if sku_key not in vistos:
                vistos.add(sku_key)
                todos.append(pr)
                total_nuevos += 1
            if limite and len(todos) >= limite: break

        if limite and len(todos) >= limite: break
        if p_num >= max_paginas: break
        p_num += 1

    log(f"  [{id_cat}] {cat_nombre}: {total_nuevos} nuevos (pág {p_num})")

# ─── MODO RÁPIDO ─────────────────────────────────────────────────────────────
async def modo_rapido(page, cats_activas):
    """
    Lee el último JSON completo, recorre solo los listados de las categorías
    activas y detecta cambios de stock y precio. Genera stock_delta_FECHA.json
    con SOLO los SKUs que cambiaron respecto al JSON anterior.
    Tiempo: ~3 minutos para calzado + ropa.
    """
    log("=== MODO RÁPIDO: actualizando stock y precios ===")

    json_anterior = ultimo_json_completo()
    if not json_anterior:
        log("ERROR: No hay JSON completo previo en output_vicsa/")
        log("  Corre primero: python scraper_v13.py (modo completo)")
        sys.exit(1)

    log(f"Referencia: {json_anterior.name}")

    with open(json_anterior, encoding="utf-8") as f:
        data_anterior = json.load(f)

    # Construir mapa sku → {precio, stock_publicar} del JSON anterior
    sku_anterior = {}
    for prod in data_anterior.get("productos", []):
        for v in prod.get("variantes", []):
            sku_anterior[v["sku"]] = {
                "precio_neto":   v["precio_neto"],
                "stock_publicar": v.get("stock_publicar", 1),
            }

    log(f"SKUs en referencia: {len(sku_anterior)}")

    # Recorrer listados rápido (sin visitar fichas individuales)
    todos  = []
    vistos = set()
    cats_v = set()

    for id_cat, cat_nombre in cats_activas.items():
        if id_cat in cats_v: continue
        cats_v.add(id_cat)
        await scrape_categoria(page, id_cat, cat_nombre, 0, todos, vistos)

    # Detectar cambios
    cambios = []
    nuevos  = []
    sin_stock_ahora = []

    sku_actual = {p["sku"]: p for p in todos}

    for sku, datos_act in sku_actual.items():
        if sku not in sku_anterior:
            nuevos.append(sku)
            cambios.append({**datos_act, "cambio": "nuevo"})
            continue

        datos_ant = sku_anterior[sku]
        precio_cambio = abs(datos_act["precio_neto"] - datos_ant["precio_neto"]) > 1
        stock_cambio  = datos_act["stock_publicar"] != datos_ant["stock_publicar"]

        if datos_act["stock_publicar"] == 0:
            sin_stock_ahora.append(sku)

        if precio_cambio or stock_cambio:
            cambios.append({
                **datos_act,
                "cambio":           "actualizado",
                "precio_anterior":  datos_ant["precio_neto"],
                "stock_anterior":   datos_ant["stock_publicar"],
            })

    # SKUs que estaban en el JSON anterior y ya no aparecen = sin stock
    for sku in sku_anterior:
        if sku not in sku_actual:
            cambios.append({
                "sku":            sku,
                "stock_vicsa":    0,
                "stock_publicar": 0,
                "precio_neto":    sku_anterior[sku]["precio_neto"],
                "cambio":         "desaparecido",
            })
            sin_stock_ahora.append(sku)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    delta_path = OUTPUT_DIR / f"stock_delta_{ts}.json"

    resultado = {
        "generado":         datetime.now().isoformat(),
        "modo":             "rapido",
        "json_referencia":  json_anterior.name,
        "total_skus_vistos": len(todos),
        "total_cambios":    len(cambios),
        "nuevos":           len(nuevos),
        "sin_stock":        len(sin_stock_ahora),
        "cambios":          cambios,
    }

    with open(delta_path, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    log(f"\n{'='*55}")
    log(f"MODO RÁPIDO COMPLETO")
    log(f"  SKUs visitados: {len(todos)}")
    log(f"  Cambios detectados: {len(cambios)}")
    log(f"    - Nuevos: {len(nuevos)}")
    log(f"    - Sin stock: {len(sin_stock_ahora)}")
    log(f"    - Precio/stock actualizado: {len(cambios) - len(nuevos)}")
    log(f"  Archivo delta: {delta_path}")
    log(f"{'='*55}")

    return delta_path

# ─── Guardado (modo completo) ────────────────────────────────────────────────
def guardar_json(productos, path):
    """
    JSON estructurado por producto base, agrupando variantes.
    DEDUPLICACIÓN: cada SKU aparece una sola vez por producto base.
    Incluye stock_publicar (conservador) para sync_ml.py
    """
    grupos = {}
    for p in productos:
        nombre_base = re.sub(
            r'\s+(XXXL|XXL|XL|XS|XXS|\bL\b|\bM\b|\bS\b|T\.\d{2}|TALLA\s+\S+|\d{2}|UNICA).*$',
            '', p["nombre"], flags=re.IGNORECASE
        ).strip()
        nombre_base = re.sub(
            r'\s+(' + '|'.join(c.upper() for c in COLORES) + r').*$',
            '', nombre_base, flags=re.IGNORECASE
        ).strip()

        key = f"{p['categoria']}||{nombre_base}"
        if key not in grupos:
            grupos[key] = {
                "nombre_base":      nombre_base,
                "categoria":        p["categoria"],
                "especificaciones": p["especificaciones"],
                "ficha_tecnica_url":p["ficha_tecnica_url"],
                "certificaciones":  p["certificaciones"],
                "imagen_principal": p["imagen"],
                "variantes":        [],
                "_skus_vistos":     set(),  # para deduplicar
            }

        sku = p["sku"]
        if sku in grupos[key]["_skus_vistos"]:
            continue  # ← FIX: evita duplicados del mismo SKU
        grupos[key]["_skus_vistos"].add(sku)

        grupos[key]["variantes"].append({
            "sku":           sku,
            "talla":         p["talla"],
            "color":         p["color"],
            "precio_neto":   p["precio_neto"],
            "stock_vicsa":   p.get("stock_vicsa", 5),
            "stock_publicar":p.get("stock_publicar", 1),
            "imagen":        p["imagen"],
            "url":           p["url_producto"],
            "actualizado":   p["_actualizado"],
        })

    # Limpiar set interno antes de guardar
    productos_limpios = []
    for g in grupos.values():
        g.pop("_skus_vistos", None)
        productos_limpios.append(g)

    resultado = {
        "generado":        datetime.now().isoformat(),
        "modo":            "completo",
        "total_skus":      len(productos),
        "total_productos": len(productos_limpios),
        "productos":       productos_limpios,
    }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)

    total_vars = sum(len(g["variantes"]) for g in productos_limpios)
    log(f"JSON: {path} ({len(productos)} SKUs raw → {total_vars} únicos → {len(productos_limpios)} productos base)")

def guardar_csv(productos, path):
    campos = ["sku","nombre","categoria","talla","color","precio_neto",
              "stock_vicsa","stock_publicar",
              "especificaciones","ficha_tecnica_url","certificaciones","imagen","url_producto"]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=campos, extrasaction="ignore")
        w.writeheader()
        w.writerows(productos)
    log(f"CSV: {path} ({len(productos)} SKUs)")

def guardar_xlsx(productos, path):
    if not TIENE_XLSX: return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CATALOGO"
    ws.sheet_view.showGridLines = False
    NA="E85D04"; BL="FFFFFF"; GR="1A1A1A"; AL="F7F5F2"
    def fill(h): return PatternFill("solid", start_color=h, fgColor=h)
    def thin(): return Border(bottom=Side(style="thin", color="DDDDDD"))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    ws.merge_cells("A1:M1")
    ws["A1"] = f"Catalogo VICSA - RAI SPA  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = Font(name="Arial", bold=True, color=BL, size=13)
    ws["A1"].fill = fill(NA)
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 30
    hdrs = [("SKU",16),("NOMBRE",45),("CATEGORÍA",22),("TALLA",10),("COLOR",14),
            ("PRECIO COSTO $",16),("STOCK VICSA",12),("STOCK ML",10),
            ("ESPECIFICACIONES",45),("FICHA URL",35),("CERTIF.",35),
            ("IMAGEN",22),("URL PRODUCTO",40)]
    for c,(h,w) in enumerate(hdrs,1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, color=BL, size=9)
        cell.fill = fill(GR)
        cell.alignment = center
        ws.column_dimensions[ws.cell(row=2,column=c).column_letter].width = w
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"
    campos = ["sku","nombre","categoria","talla","color","precio_neto",
              "stock_vicsa","stock_publicar",
              "especificaciones","ficha_tecnica_url","certificaciones","imagen","url_producto"]
    for i,p in enumerate(productos):
        r = i+3
        bg = BL if i%2==0 else AL
        for c,campo in enumerate(campos,1):
            val = p.get(campo,"")
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = fill(bg)
            cell.border = thin()
            es_precio = (c==6)
            es_stock0 = (c==8 and val==0)
            cell.font = Font(
                name="Arial", size=10,
                color=("FF0000" if es_stock0 else (NA if es_precio else "111111")),
                bold=(es_precio or es_stock0)
            )
            cell.alignment = left
            if c==6: cell.number_format = "$#,##0"
        ws.row_dimensions[r].height = 36
    wb.save(path)
    log(f"Excel: {path} ({len(productos)} SKUs)")

# ─── Main ─────────────────────────────────────────────────────────────────────
async def main():
    parser = argparse.ArgumentParser(description="VICSA Scraper v13 - Dos velocidades")
    parser.add_argument("--modo",         choices=["rapido","completo"], default="completo",
                        help="rapido: solo stock/precio (~3 min) | completo: catálogo completo (~30 min)")
    parser.add_argument("--limite",       type=int, default=0,
                        help="Limitar número total de SKUs (0 = sin límite, útil para pruebas)")
    parser.add_argument("--sin-imagenes", action="store_true",
                        help="No descargar imágenes (modo completo más rápido)")
    parser.add_argument("--sin-detalle",  action="store_true",
                        help="No visitar fichas individuales (modo completo más rápido)")
    parser.add_argument("--categorias",   nargs="+", default=[],
                        help='Filtrar categorías. Ej: "Calzado de Seguridad" "Ropa Tecnica"')
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(exist_ok=True)
    if not args.sin_imagenes and args.modo == "completo":
        IMG_DIR.mkdir(exist_ok=True)

    # Filtrar categorías
    cats_activas = CATEGORIAS
    if args.categorias:
        cats_norm = [c.lower().strip() for c in args.categorias]
        cats_activas = {k: v for k, v in CATEGORIAS.items() if v.lower() in cats_norm}
        if not cats_activas:
            print(f"ERROR: Ninguna categoría coincide con: {args.categorias}")
            print(f"Disponibles: {sorted(set(CATEGORIAS.values()))}")
            sys.exit(1)
        log(f"Categorías activas: {sorted(set(cats_activas.values()))}")

    # ── MODO RÁPIDO ──────────────────────────────────────────────────────────
    if args.modo == "rapido":
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True, args=["--no-sandbox","--disable-dev-shm-usage"])
            ctx  = await browser.new_context(locale="es-CL", viewport={"width":1280,"height":900})
            page = await ctx.new_page()
            await login_automatico(page)
            await modo_rapido(page, cats_activas)
            await browser.close()
        return

    # ── MODO COMPLETO ─────────────────────────────────────────────────────────
    log(f"\nFase 1: Recorriendo {len(cats_activas)} categorías...")
    CHECKPOINT = OUTPUT_DIR / "checkpoint_progreso.json"
    todos  = []
    vistos = set()
    cats_v = set()

    # Reanudar desde checkpoint si existe
    if CHECKPOINT.exists():
        try:
            with open(CHECKPOINT, encoding="utf-8") as f:
                cp = json.load(f)
            todos  = cp.get("productos", [])
            vistos = set(cp.get("vistos", []))
            cats_v = set(cp.get("cats_completadas", []))
            log(f"Checkpoint cargado: {len(todos)} SKUs, {len(cats_v)} cats completadas — reanudando...")
        except Exception:
            log("Checkpoint corrupto, iniciando desde cero")
            todos, vistos, cats_v = [], set(), set()

    for id_cat, cat_nombre in list(cats_activas.items()):
        if args.limite and len(todos) >= args.limite: break
        if id_cat in cats_v: continue

        # Categorías que crashean el browser — saltar y marcar como completadas
        CATS_SKIP = {"836"}
        if id_cat in CATS_SKIP:
            log(f"  [{id_cat}] {cat_nombre}: saltando (categoría problemática)")
            cats_v.add(id_cat)
            continue

        # Cada categoría abre su propio contexto de playwright limpio
        intentos = 0
        while intentos < 3:
            try:
                async with async_playwright() as pw:
                    browser = await pw.chromium.launch(
                        headless=True, args=["--no-sandbox","--disable-dev-shm-usage"])
                    ctx  = await browser.new_context(locale="es-CL", viewport={"width":1280,"height":900})
                    page = await ctx.new_page()
                    await login_automatico(page)
                    await scrape_categoria(page, id_cat, cat_nombre, args.limite, todos, vistos)
                    await browser.close()
                cats_v.add(id_cat)
                break  # éxito, salir del while
            except Exception as e:
                intentos += 1
                log(f"Error cat {id_cat} (intento {intentos}/3): {str(e)[:80]}")
                await asyncio.sleep(3)
                if intentos == 3:
                    log(f"  Saltando categoría {id_cat} después de 3 intentos")

        # Checkpoint después de cada categoría
        with open(CHECKPOINT, "w", encoding="utf-8") as f:
            json.dump({
                "productos":        todos,
                "vistos":           list(vistos),
                "cats_completadas": list(cats_v),
                "guardado":         datetime.now().isoformat(),
            }, f, ensure_ascii=False)

    log(f"Total SKUs fase 1: {len(todos)}")

    if not args.sin_detalle:
        log("\nFase 2: Enriqueciendo fichas individuales...")
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=True, args=["--no-sandbox","--disable-dev-shm-usage"])
            ctx  = await browser.new_context(locale="es-CL", viewport={"width":1280,"height":900})
            page = await ctx.new_page()
            await login_automatico(page)
            for i, prod in enumerate(todos):
                if i % 20 == 0:
                    log(f"  [{i+1}/{len(todos)}] {prod['nombre'][:50]}")
                try:
                    await enriquecer_producto(page, prod, not args.sin_imagenes)
                except Exception:
                    pass
            await browser.close()

    # Eliminar checkpoint al terminar exitosamente
    if CHECKPOINT.exists():
        CHECKPOINT.unlink()

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    guardar_json(todos,  OUTPUT_DIR / f"catalogo_vicsa_{ts}.json")
    guardar_csv(todos,   OUTPUT_DIR / f"catalogo_vicsa_{ts}.csv")
    guardar_xlsx(todos,  OUTPUT_DIR / f"catalogo_vicsa_{ts}.xlsx")

    n_imgs = len(list(IMG_DIR.glob("*"))) if IMG_DIR.exists() else 0
    log(f"\n{'='*55}")
    log(f"MODO COMPLETO LISTO")
    log(f"  SKUs procesados: {len(todos)}")
    log(f"  Imágenes:        {n_imgs}")
    log(f"  Archivos en:     {OUTPUT_DIR.resolve()}")
    log(f"{'='*55}")

if __name__ == "__main__":
    asyncio.run(main())


# ═══════════════════════════════════════════════════════════════════════════
# GUÍA DE USO v13
# ═══════════════════════════════════════════════════════════════════════════
#
# ── INSTALACIÓN (una sola vez) ───────────────────────────────────────────
#   pip install playwright openpyxl
#   playwright install chromium
#
# ── VARIABLES DE ENTORNO (PowerShell) ───────────────────────────────────
#   $env:VICSA_USER = "cristobalcontrerasc@gmail.com"
#   $env:VICSA_PASS = "tu_contraseña"
#
# ── MODO COMPLETO — piloto calzado + ropa (~30 min) ─────────────────────
#   python scraper_v13.py --categorias "Calzado de Seguridad" "Ropa Tecnica" "Ropa de Trabajo"
#
# ── MODO RÁPIDO — solo stock/precio (~3 min) ────────────────────────────
#   python scraper_v13.py --modo rapido --categorias "Calzado de Seguridad" "Ropa Tecnica" "Ropa de Trabajo"
#
# ── PRUEBA RÁPIDA — 10 productos, sin detalles ──────────────────────────
#   python scraper_v13.py --limite 10 --sin-detalle --sin-imagenes --categorias "Calzado de Seguridad"
#
# ── PRODUCCIÓN COMPLETA (todos los productos) ───────────────────────────
#   python scraper_v13.py
#
# ── ARCHIVOS GENERADOS ──────────────────────────────────────────────────
#   output_vicsa/
#   ├── catalogo_vicsa_FECHA.json   ← modo completo → lo lee sync_ml.py
#   ├── catalogo_vicsa_FECHA.csv
#   ├── catalogo_vicsa_FECHA.xlsx
#   ├── stock_delta_FECHA.json      ← modo rápido → solo cambios → sync_ml.py
#   └── imagenes/                   ← imágenes de productos
#
# ── GITHUB ACTIONS (sync.yml — próximo paso) ────────────────────────────
#   Modo rápido:   cron '0 * * * *'   → cada hora
#   Modo completo: cron '0 3 * * *'   → cada día a las 3am
#
# ═══════════════════════════════════════════════════════════════════════════
